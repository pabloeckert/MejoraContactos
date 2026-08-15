"""Genera la lista maestra a partir de la vista calculada sobre clusters
(el "contacto maestro" no existe como fila propia, se materializa uniendo
todos los raw_records que comparten cluster_id). Nunca toca Data/Crudos/.

Salida en .xlsx nativo (no CSV — pedido explícito: abrir un CSV con
caracteres acentuados directo en Excel en Windows suele mostrar la
codificación rota si no lleva BOM; un .xlsx real vía openpyxl no tiene ese
problema porque no pasa por texto plano en ningún momento).

Un WhatsApp/Teléfono fijo/Email por celda, nunca varios juntos separados
por ";" — pedido explícito. Si un contacto tiene más de un valor en
cualquiera de esos tres campos, se generan varias filas (misma persona,
un valor distinto por fila) en vez de amontonarlos en un solo casillero.

Esquema de columnas fijado en la Ficha 15.1 de la encuesta de cierre:
nombre, apellido, cargo, empresa, whatsapp/telefono_fijo por separado,
tag (etiqueta, hoy vacía — auto-etiquetado queda para una fase posterior),
domicilio/ciudad/provincia/pais opcionales, y nota_referencia (reutiliza
el campo "notas" de origen). Sin columna "fuentes" — no estaba pedida."""

from __future__ import annotations

import json
import re
import sqlite3
from itertools import zip_longest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from motor.config import Config
from motor.email_normalizer import normalizar_campo_email
from motor.phone_normalizer import normalizar_campo_telefono

_COLUMNAS = [
    "nombre",
    "apellido",
    "cargo",
    "empresa",
    "whatsapp",
    "telefono_fijo",
    "email",
    "tag",
    "domicilio",
    "ciudad",
    "provincia",
    "pais",
    "cumpleanos",
    "foto_url",
    "nota_referencia",
]

_ENCABEZADOS = {
    "nombre": "Nombre",
    "apellido": "Apellido",
    "cargo": "Cargo",
    "empresa": "Empresa",
    "whatsapp": "WhatsApp",
    "telefono_fijo": "Teléfono fijo",
    "email": "Email",
    "tag": "Tag",
    "domicilio": "Domicilio",
    "ciudad": "Ciudad",
    "provincia": "Provincia",
    "pais": "País",
    "cumpleanos": "Cumpleaños",
    "foto_url": "Foto",
    "nota_referencia": "Nota de referencia",
}

_ANCHO_COLUMNA = {
    "nombre": 20,
    "apellido": 20,
    "cargo": 22,
    "empresa": 24,
    "whatsapp": 20,
    "telefono_fijo": 20,
    "email": 30,
    "tag": 14,
    "domicilio": 26,
    "ciudad": 16,
    "provincia": 16,
    "pais": 14,
    "cumpleanos": 14,
    "foto_url": 30,
    "nota_referencia": 34,
}


def exportar_lista_maestra(config: Config, conn: sqlite3.Connection) -> Path:
    clusters = _materializar_clusters(conn)
    _aplicar_ediciones_manuales(conn, clusters)
    filas: list[dict] = []
    for cluster in clusters:
        filas.extend(_expandir_filas(cluster))
    filas.sort(key=lambda f: (f["apellido"].lower(), f["nombre"].lower()))

    destino = config.rutas.carpeta_salida / "lista-maestra.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista maestra"

    ws.append([_ENCABEZADOS[c] for c in _COLUMNAS])
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1A3D84")
        celda.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNAS))}1"

    for fila in filas:
        ws.append([fila[c] for c in _COLUMNAS])

    for idx, columna in enumerate(_COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _ANCHO_COLUMNA[columna]

    wb.save(destino)
    return destino


def _aplicar_ediciones_manuales(conn: sqlite3.Connection, clusters: list[dict]) -> None:
    """Una edición manual guardada desde el revisor web (/editar/<cluster_id>)
    pisa lo que haya calculado la limpieza automática para ESE cluster —
    la corrección humana siempre gana. No toca raw_records ni
    normalized_records, así que sigue sin haber nada destructivo: si se
    borra la edición manual, el cluster vuelve a mostrar el valor
    calculado."""
    filas = conn.execute("SELECT * FROM ediciones_manuales").fetchall()
    ediciones = {fila["cluster_id"]: fila for fila in filas}

    _CAMPOS = ("nombre", "apellido", "cargo", "tag", "domicilio", "ciudad", "provincia", "pais")
    _CAMPOS_MULTIVALOR = {
        "whatsapp_json": "whatsapp",
        "telefono_fijo_json": "telefono_fijo",
        "emails_json": "emails",
    }
    for cluster in clusters:
        edicion = ediciones.get(cluster["cluster_id"])
        if not edicion:
            continue
        cluster["editado_manualmente"] = True
        for campo in _CAMPOS:
            valor = edicion[campo]
            if valor is not None and valor != "":
                cluster[campo] = valor
        if edicion["organizacion"] is not None and edicion["organizacion"] != "":
            cluster["organizacion"] = edicion["organizacion"]
        if edicion["notas"] is not None and edicion["notas"] != "":
            cluster["nota_referencia"] = edicion["notas"]
        for columna, campo_cluster in _CAMPOS_MULTIVALOR.items():
            valor = edicion[columna] if columna in edicion.keys() else None
            if valor is not None:
                cluster[campo_cluster] = set(json.loads(valor))


def listar_contactos(conn: sqlite3.Connection, pagina: int = 1, tamano: int = 100) -> tuple[list[dict], int]:
    """Todos los clusters materializados (con ediciones manuales ya
    aplicadas), ordenados igual que el export, para listar/paginar desde la
    API JSON (Fase 1 — UI nueva) sin tener que generar el .xlsx completo."""
    clusters = _materializar_clusters(conn)
    _aplicar_ediciones_manuales(conn, clusters)
    clusters.sort(key=lambda c: (c["apellido"].lower(), c["nombre"].lower()))
    total = len(clusters)
    inicio = (pagina - 1) * tamano
    return clusters[inicio : inicio + tamano], total


def exportar_whatsapp_csv(config: Config, conn: sqlite3.Connection) -> Path:
    """CSV en el formato exacto que espera MejoraWS
    (C:\\Github\\Herramientas\\MejoraWS, "Importar CSV/Excel"): columnas
    nombre,telefono,variable -- teléfono en E.164 SIN el "+" ("código de
    país, sin espacios ni signos", así lo pide su propio README). Un
    contacto con más de un WhatsApp genera una fila por número, mismo
    criterio que la lista maestra. La columna "variable" lleva el tag
    (familiar/laboral/cliente/proveedor/personal) -- útil como variable de
    personalización del mensaje ({variable}) si el usuario quiere, no es
    obligatorio usarla."""
    import csv

    clusters = _materializar_clusters(conn)
    _aplicar_ediciones_manuales(conn, clusters)

    destino = config.rutas.carpeta_salida / "contactos-whatsapp.csv"
    destino.parent.mkdir(parents=True, exist_ok=True)

    filas: list[tuple[str, str, str]] = []
    for cluster in clusters:
        nombre_completo = f"{cluster['nombre']} {cluster['apellido']}".strip()
        if not nombre_completo:
            continue
        for whatsapp in sorted(cluster["whatsapp"]):
            filas.append((nombre_completo, whatsapp.lstrip("+"), cluster["tag"] or ""))

    filas.sort(key=lambda f: f[0].lower())

    with destino.open("w", newline="", encoding="utf-8") as f:
        escritor = csv.writer(f)
        escritor.writerow(["nombre", "telefono", "variable"])
        escritor.writerows(filas)

    return destino


def obtener_contacto(conn: sqlite3.Connection, cluster_id: str) -> dict | None:
    """Un cluster materializado (con ediciones manuales ya aplicadas),
    para precargar el formulario de /editar en el revisor web."""
    clusters = [c for c in _materializar_clusters(conn) if c["cluster_id"] == cluster_id]
    if not clusters:
        return None
    _aplicar_ediciones_manuales(conn, clusters)
    return clusters[0]


def buscar_contactos(conn: sqlite3.Connection, consulta: str, limite: int = 30) -> list[dict]:
    """Busca por nombre/apellido/organización/teléfono/email (FTS5, prefijo
    por palabra) y devuelve un cluster materializado por cada contacto que
    matchea — para elegir a quién editar desde /buscar."""
    consulta_fts = _sanear_consulta_fts(consulta)
    if not consulta_fts:
        return []

    filas = conn.execute(
        "SELECT n.raw_record_id FROM busqueda_fts f "
        "JOIN normalized_records n ON n.id = f.rowid "
        "WHERE busqueda_fts MATCH ? LIMIT ?",
        (consulta_fts, limite),
    ).fetchall()
    if not filas:
        return []

    raw_ids = [f["raw_record_id"] for f in filas]
    marcadores = ",".join("?" * len(raw_ids))
    cluster_ids = {
        f["cluster_id"]
        for f in conn.execute(
            f"SELECT DISTINCT cluster_id FROM clusters WHERE raw_record_id IN ({marcadores})",
            raw_ids,
        ).fetchall()
    }
    clusters = [c for c in _materializar_clusters(conn) if c["cluster_id"] in cluster_ids]
    _aplicar_ediciones_manuales(conn, clusters)
    return clusters


def _sanear_consulta_fts(consulta: str) -> str:
    # FTS5 rompe con caracteres sueltos como "+" o "@" en la sintaxis de
    # MATCH — se arma una consulta de prefijo por palabra ("juan* perez*")
    # a partir de los tokens alfanuméricos, nunca se manda el texto crudo.
    tokens = re.findall(r"\w+", consulta, re.UNICODE)
    return " ".join(f"{t}*" for t in tokens)


def guardar_edicion_manual(
    conn: sqlite3.Connection, cluster_id: str, campos: dict[str, str], config: Config
) -> None:
    from datetime import datetime, timezone

    valores = {
        campo: (campos.get(campo) or "").strip() or None
        for campo in ("nombre", "apellido", "cargo", "organizacion", "tag", "domicilio", "ciudad", "provincia", "pais", "notas")
    }

    existente = conn.execute(
        "SELECT whatsapp_json, telefono_fijo_json, emails_json FROM ediciones_manuales WHERE cluster_id = ?",
        (cluster_id,),
    ).fetchone()

    multivalor = {
        "whatsapp_json": _normalizar_multivalor_edicion(
            campos.get("whatsapp"),
            lambda linea: [
                r.e164 for r in normalizar_campo_telefono(linea, config.telefono, "whatsapp") if r.e164
            ],
            existente["whatsapp_json"] if existente else None,
        ),
        "telefono_fijo_json": _normalizar_multivalor_edicion(
            campos.get("telefono_fijo"),
            lambda linea: [
                r.e164 for r in normalizar_campo_telefono(linea, config.telefono, "fijo") if r.e164
            ],
            existente["telefono_fijo_json"] if existente else None,
        ),
        "emails_json": _normalizar_multivalor_edicion(
            campos.get("email"),
            lambda linea: [r.normalizado for r in normalizar_campo_email(linea, config.email) if r.normalizado],
            existente["emails_json"] if existente else None,
        ),
    }

    conn.execute(
        "INSERT INTO ediciones_manuales "
        "(cluster_id, nombre, apellido, cargo, organizacion, tag, domicilio, ciudad, provincia, pais, notas, "
        "whatsapp_json, telefono_fijo_json, emails_json, actualizado_en) "
        "VALUES (:cluster_id, :nombre, :apellido, :cargo, :organizacion, :tag, :domicilio, :ciudad, :provincia, :pais, :notas, "
        ":whatsapp_json, :telefono_fijo_json, :emails_json, :actualizado_en) "
        "ON CONFLICT(cluster_id) DO UPDATE SET "
        "nombre=excluded.nombre, apellido=excluded.apellido, cargo=excluded.cargo, "
        "organizacion=excluded.organizacion, tag=excluded.tag, domicilio=excluded.domicilio, "
        "ciudad=excluded.ciudad, provincia=excluded.provincia, pais=excluded.pais, "
        "notas=excluded.notas, whatsapp_json=excluded.whatsapp_json, "
        "telefono_fijo_json=excluded.telefono_fijo_json, emails_json=excluded.emails_json, "
        "actualizado_en=excluded.actualizado_en",
        {
            **valores,
            "cluster_id": cluster_id,
            **multivalor,
            "actualizado_en": datetime.now(timezone.utc).isoformat(),
        },
    )
    conn.commit()


def _normalizar_multivalor_edicion(raw, transformar, valor_previo):
    """Una línea por valor en el textarea del panel; cada línea se normaliza
    con la misma lógica que el resto del pipeline (phone_normalizer /
    email_normalizer) — nunca se guarda un valor roto. Campo vacío = borrar
    la corrección manual (NULL, vuelve a mostrar el valor calculado). Si el
    usuario tipeó algo pero ninguna línea normalizó a un valor válido, no se
    pisa lo que ya había guardado — mejor no guardar nada que guardar una
    lista vacía y borrar un WhatsApp/email real por un typo."""
    if raw is None or not raw.strip():
        return None
    valores: list[str] = []
    for linea in raw.splitlines():
        if linea.strip():
            valores.extend(transformar(linea))
    if not valores:
        return valor_previo
    return json.dumps(sorted(set(valores)))


def _expandir_filas(cluster: dict) -> list[dict]:
    """Un WhatsApp/Teléfono fijo/Email por fila. Si el contacto tiene más
    de un valor en cualquiera de los tres, se generan varias filas —
    alineadas por posición entre las tres listas (no hay una
    correspondencia "real" entre el 2do teléfono y el 2do email de un
    mismo contacto, es solo para no perder ningún valor sin amontonar
    todo en un casillero)."""
    whatsapp = sorted(cluster["whatsapp"]) or [""]
    telefono_fijo = sorted(cluster["telefono_fijo"]) or [""]
    emails = sorted(cluster["emails"]) or [""]

    filas = []
    for wa, fijo, mail in zip_longest(whatsapp, telefono_fijo, emails, fillvalue=""):
        filas.append(
            {
                "nombre": cluster["nombre"],
                "apellido": cluster["apellido"],
                "cargo": cluster["cargo"],
                "empresa": cluster["organizacion"],
                "whatsapp": wa,
                "telefono_fijo": fijo,
                "email": mail,
                "tag": cluster["tag"],
                "domicilio": cluster["domicilio"],
                "ciudad": cluster["ciudad"],
                "provincia": cluster["provincia"],
                "pais": cluster["pais"],
                "cumpleanos": cluster["cumpleanos"],
                "foto_url": cluster["foto_url"],
                "nota_referencia": cluster["nota_referencia"],
            }
        )
    return filas


def _materializar_clusters(conn: sqlite3.Connection) -> list[dict]:
    filas = conn.execute(
        "SELECT c.cluster_id, n.nombre, n.apellido, n.organizacion, n.cargo, "
        "n.telefonos_e164, n.telefonos_fijo_e164, n.emails, n.tag, "
        "n.domicilio, n.ciudad, n.provincia, n.pais, n.cumpleanos, n.foto_url, n.notas, n.flags "
        "FROM clusters c "
        "JOIN raw_records r ON r.id = c.raw_record_id "
        "JOIN normalized_records n ON n.raw_record_id = r.id"
    ).fetchall()

    por_cluster: dict[str, dict] = {}
    for fila in filas:
        cluster = por_cluster.setdefault(
            fila["cluster_id"],
            {
                "cluster_id": fila["cluster_id"],
                "nombre": "",
                "apellido": "",
                "cargo": "",
                "organizacion": "",
                "whatsapp": set(),
                "telefono_fijo": set(),
                "emails": set(),
                "tag": "",
                "domicilio": "",
                "ciudad": "",
                "provincia": "",
                "pais": "",
                "cumpleanos": "",
                "foto_url": "",
                "notas": [],
                # Unión de los flags de cada normalized_record que compone
                # este cluster (ej. "telefono:movil-asumido",
                # "telefono:incompleto") -- ya se calculaban en
                # normalize_pipeline.py pero se perdían acá, nunca llegaban
                # a la lista maestra ni a la UI (hallazgo real de la
                # revisión UX del 2026-08-15: el sistema sabía que un dato
                # era una suposición/corrección, pero no lo mostraba).
                "flags": set(),
                "editado_manualmente": False,
            },
        )
        cluster["nombre"] = cluster["nombre"] or fila["nombre"] or ""
        cluster["apellido"] = cluster["apellido"] or fila["apellido"] or ""
        cluster["cargo"] = cluster["cargo"] or fila["cargo"] or ""
        cluster["organizacion"] = cluster["organizacion"] or fila["organizacion"] or ""
        cluster["whatsapp"] |= set(json.loads(fila["telefonos_e164"]))
        cluster["telefono_fijo"] |= set(json.loads(fila["telefonos_fijo_e164"]))
        cluster["emails"] |= set(json.loads(fila["emails"]))
        cluster["tag"] = cluster["tag"] or fila["tag"] or ""
        cluster["domicilio"] = cluster["domicilio"] or fila["domicilio"] or ""
        cluster["ciudad"] = cluster["ciudad"] or fila["ciudad"] or ""
        cluster["provincia"] = cluster["provincia"] or fila["provincia"] or ""
        cluster["pais"] = cluster["pais"] or fila["pais"] or ""
        cluster["cumpleanos"] = cluster["cumpleanos"] or fila["cumpleanos"] or ""
        cluster["foto_url"] = cluster["foto_url"] or fila["foto_url"] or ""
        cluster["flags"] |= set(json.loads(fila["flags"] or "[]"))
        if fila["notas"]:
            cluster["notas"].append(fila["notas"])

    for cluster in por_cluster.values():
        cluster["nota_referencia"] = " | ".join(dict.fromkeys(cluster["notas"]))

    return list(por_cluster.values())
