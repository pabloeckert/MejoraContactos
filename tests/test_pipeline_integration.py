"""Prueba end-to-end del pipeline (extraer -> normalizar -> deduplicar ->
exportar -> deshacer) sobre archivos CSV sintéticos, nunca sobre
pablo.csv/Sindy.csv reales. El LLM-judge queda desactivado (llm.activar_
para_dudosos=False) para no depender de red ni de API keys en los tests;
eso empuja los casos ambiguos directo a revision_pendiente, que también se
cubre acá."""

from openpyxl import load_workbook

from motor.config import Config, DedupConfig, EmailConfig, LlmConfig, RevisorConfig, RutasConfig, TelefonoConfig
from motor.dedup.merge_engine import aplicar_decision_lote, deduplicar_todo, deshacer, deshacer_ultima_corrida
from motor.export import exportar_lista_maestra, exportar_whatsapp_csv, guardar_edicion_manual
from motor.ingest import extraer_todo
from motor.normalize_pipeline import normalizar_todo
from motor.staging_db import conectar


def _config_prueba(tmp_path):
    (tmp_path / "Crudos").mkdir()
    return Config(
        rutas=RutasConfig(
            carpeta_raiz=tmp_path / "Crudos",
            carpeta_salida=tmp_path / "Salida",
            base_sqlite=tmp_path / "Salida" / "staging.sqlite",
        ),
        extensiones_permitidas=frozenset({"csv"}),
        telefono=TelefonoConfig(),
        email=EmailConfig(),
        dedup=DedupConfig(),
        llm=LlmConfig(activar_para_dudosos=False),
        revisor=RevisorConfig(),
    )


def test_pipeline_completo_fusiona_por_telefono_y_exporta(tmp_path):
    config = _config_prueba(tmp_path)

    # Dos archivos de origen distinto (simula pablo.csv / Sindy.csv) con
    # un contacto repetido por teléfono, y uno sin relación.
    (config.rutas.carpeta_raiz / "pablo.csv").write_text(
        "Nombre,Apellido,Telefono,Correo\nJuan,Perez,3743504517,juan@gmail.com\n",
        encoding="utf-8",
    )
    (config.rutas.carpeta_raiz / "sindy.csv").write_text(
        "Nombre,Apellido,Telefono\nJ,P,3743504517\nRicardo,Gomez,3764368724\n",
        encoding="utf-8",
    )

    conn = conectar(config.rutas.base_sqlite)

    assert extraer_todo(config, conn) == 3
    assert normalizar_todo(config, conn) == 3

    resultado = deduplicar_todo(config, conn)
    assert resultado["regla"] >= 1  # el par Juan/J se fusiona por teléfono exacto

    destino = exportar_lista_maestra(config, conn)
    assert destino.suffix == ".xlsx"
    filas = _leer_xlsx(destino)
    # 2 personas reales: el duplicado fusionado + Ricardo (sin relación)
    assert len(filas) == 2
    assert any("juan@gmail.com" in fila["Email"] for fila in filas)


def _leer_xlsx(path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    encabezados = [c.value for c in ws[1]]
    return [
        dict(zip(encabezados, [c.value or "" for c in fila]))
        for fila in ws.iter_rows(min_row=2)
    ]


def test_correr_extraer_dos_veces_sin_cambios_no_duplica(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "a.csv").write_text("Nombre,Telefono\nAna,111111\n", encoding="utf-8")

    conn = conectar(config.rutas.base_sqlite)
    assert extraer_todo(config, conn) == 1
    assert extraer_todo(config, conn) == 0  # nada nuevo, mismo hash


def test_deshacer_revierte_una_fusion(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "dup.csv").write_text(
        "Nombre,Telefono\nJuan,3743504517\nJ,3743504517\n", encoding="utf-8"
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    fila = conn.execute("SELECT cluster_id FROM clusters LIMIT 1").fetchone()
    cluster_id = fila["cluster_id"]
    miembros_antes = conn.execute(
        "SELECT raw_record_id FROM clusters WHERE cluster_id = ?", (cluster_id,)
    ).fetchall()
    assert len(miembros_antes) == 2  # confirma que sí se habían fusionado

    afectados = deshacer(conn, cluster_id)
    assert afectados == 2

    cluster_ids_despues = {
        fila["cluster_id"]
        for fila in conn.execute(
            "SELECT cluster_id FROM clusters WHERE raw_record_id IN (?, ?)",
            (miembros_antes[0]["raw_record_id"], miembros_antes[1]["raw_record_id"]),
        ).fetchall()
    }
    assert len(cluster_ids_despues) == 2  # ahora cada uno en su propio cluster


def test_deshacer_ultima_corrida_revierte_toda_la_corrida_de_una_vez(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "dup.csv").write_text(
        "Nombre,Telefono\nJuan,3743504517\nJ,3743504517\nAna,3764368724\nA,3764368724\n",
        encoding="utf-8",
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    resultado = deduplicar_todo(config, conn)
    assert resultado["regla"] >= 2  # dos pares fusionados por teléfono exacto

    fusionados_antes = conn.execute(
        "SELECT cluster_id, COUNT(*) c FROM clusters GROUP BY cluster_id HAVING COUNT(*) > 1"
    ).fetchall()
    assert len(fusionados_antes) == 2

    resumen = deshacer_ultima_corrida(conn)
    assert resumen["raw_records_afectados"] == 4
    assert resumen["clusters_afectados"] == 2

    fusionados_despues = conn.execute(
        "SELECT cluster_id, COUNT(*) c FROM clusters GROUP BY cluster_id HAVING COUNT(*) > 1"
    ).fetchall()
    assert fusionados_despues == []  # nadie queda fusionado


def test_deshacer_ultima_corrida_sin_corridas_no_rompe(tmp_path):
    config = _config_prueba(tmp_path)
    conn = conectar(config.rutas.base_sqlite)
    resumen = deshacer_ultima_corrida(conn)
    assert resumen == {"corrida_id": None, "clusters_afectados": 0, "raw_records_afectados": 0}


def test_export_separa_cargo_direccion_y_telefono_movil_de_fijo(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "completo.csv").write_text(
        "Nombre,Apellido,Cargo,Telefono,Address 1 - City,Address 1 - Country\n"
        "Juan,Perez,Gerente,3743504517,Posadas,Argentina\n",
        encoding="utf-8",
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)
    destino = exportar_lista_maestra(config, conn)
    assert destino.suffix == ".xlsx"

    filas = _leer_xlsx(destino)

    assert len(filas) == 1
    fila = filas[0]
    assert fila["Cargo"] == "Gerente"
    assert fila["Ciudad"] == "Posadas"
    assert fila["País"] == "Argentina"
    # sin pista de etiqueta y config.asumir_movil_por_defecto=True -> va a "WhatsApp", no a "Teléfono fijo"
    assert fila["WhatsApp"]
    assert fila["Teléfono fijo"] == ""


def test_contacto_con_dos_whatsapp_genera_dos_filas_no_una_celda_junta(tmp_path):
    config = _config_prueba(tmp_path)
    # Dos fuentes con el mismo nombre pero teléfonos distintos, sin ninguna
    # otra señal exacta en común -> no se fusionan por regla, pero sirve
    # igual para probar que un solo raw_record con 2 números en el campo
    # telefono no los junta en una celda.
    (config.rutas.carpeta_raiz / "dos_numeros.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517;3764368724\n",
        encoding="utf-8",
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)
    destino = exportar_lista_maestra(config, conn)
    filas = _leer_xlsx(destino)

    assert len(filas) == 2  # una fila por whatsapp, nunca los dos juntos en una celda
    whatsapps = {fila["WhatsApp"] for fila in filas}
    assert whatsapps == {"+5493743504517", "+5493764368724"}
    for fila in filas:
        assert ";" not in fila["WhatsApp"]
        assert fila["Nombre"] == "Juan"  # el resto del contacto se repite igual en ambas filas


def test_aplicar_decision_lote_fusiona_clusters_de_verdad_no_solo_el_log(tmp_path):
    # Caso real recurrente: mismo teléfono, nombres que el sistema lee
    # como claramente distintos (salvaguarda de scoring.py) -> cae en
    # revision_pendiente aunque el teléfono matchee exacto. Aprobar el
    # lote debe fusionar los CLUSTERS de verdad, no solo marcar el log
    # (bug real: antes de aplicar_decision_lote, el contador de
    # pendientes bajaba a 0 pero la lista maestra seguía separada).
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "compartido.csv").write_text(
        "Nombre,Apellido,Telefono\n"
        "Lucia,Fernandez,3743504517\n"
        "Gustavo,Lopez,3743504517\n",
        encoding="utf-8",
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    pendiente = conn.execute(
        "SELECT detalle FROM decisiones_log WHERE accion = 'revision_pendiente' LIMIT 1"
    ).fetchone()
    assert pendiente is not None  # confirma que el caso realmente cayó en revisión
    patron = pendiente["detalle"]

    clusters_antes = {
        fila["cluster_id"] for fila in conn.execute("SELECT cluster_id FROM clusters").fetchall()
    }
    assert len(clusters_antes) == 2  # todavía separados

    actualizados = aplicar_decision_lote(conn, patron, True)
    assert actualizados >= 1

    clusters_despues = {
        fila["cluster_id"] for fila in conn.execute("SELECT cluster_id FROM clusters").fetchall()
    }
    assert len(clusters_despues) == 1  # ahora fusionados de verdad

    pendientes_restantes = conn.execute(
        "SELECT COUNT(*) c FROM decisiones_log WHERE accion = 'revision_pendiente'"
    ).fetchone()["c"]
    assert pendientes_restantes == 0

    destino = exportar_lista_maestra(config, conn)
    filas = _leer_xlsx(destino)
    assert len(filas) == 1  # el export refleja la fusión, no dos filas separadas


def test_edicion_manual_pisa_el_valor_calculado_al_exportar(tmp_path):
    from datetime import datetime, timezone

    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "uno.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517\n", encoding="utf-8"
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    cluster_id = conn.execute("SELECT cluster_id FROM clusters LIMIT 1").fetchone()["cluster_id"]
    conn.execute(
        "INSERT INTO ediciones_manuales (cluster_id, tag, actualizado_en) VALUES (?, ?, ?)",
        (cluster_id, "familiar", datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()

    destino = exportar_lista_maestra(config, conn)
    filas = _leer_xlsx(destino)
    assert filas[0]["Tag"] == "familiar"  # la corrección manual gana sobre el auto-etiquetado


def test_edicion_manual_normaliza_whatsapp_telefono_fijo_y_email(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "uno.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517\n", encoding="utf-8"
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    cluster_id = conn.execute("SELECT cluster_id FROM clusters LIMIT 1").fetchone()["cluster_id"]
    guardar_edicion_manual(
        conn,
        cluster_id,
        {
            "whatsapp": "3764368724",
            "telefono_fijo": "3764 42-1234",
            "email": "  JUAN@GMAIL.COM  ",
        },
        config,
    )

    destino = exportar_lista_maestra(config, conn)
    filas = _leer_xlsx(destino)

    assert filas[0]["WhatsApp"] == "+5493764368724"  # normalizado a E.164, como el resto del pipeline
    assert filas[0]["Teléfono fijo"]
    assert filas[0]["Email"] == "juan@gmail.com"


def test_edicion_manual_linea_invalida_no_borra_lo_ya_guardado(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "uno.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517\n", encoding="utf-8"
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    cluster_id = conn.execute("SELECT cluster_id FROM clusters LIMIT 1").fetchone()["cluster_id"]
    guardar_edicion_manual(conn, cluster_id, {"email": "juan@gmail.com"}, config)

    # Un segundo guardado con basura (nada normaliza a un email válido) no
    # debe pisar el email bueno que ya estaba guardado.
    guardar_edicion_manual(conn, cluster_id, {"email": "###"}, config)

    fila = conn.execute(
        "SELECT emails_json FROM ediciones_manuales WHERE cluster_id = ?", (cluster_id,)
    ).fetchone()
    assert fila["emails_json"] == '["juan@gmail.com"]'


def test_edicion_manual_campo_vacio_borra_la_correccion(tmp_path):
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "uno.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517\n", encoding="utf-8"
    )

    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    cluster_id = conn.execute("SELECT cluster_id FROM clusters LIMIT 1").fetchone()["cluster_id"]
    guardar_edicion_manual(conn, cluster_id, {"email": "juan@gmail.com"}, config)
    guardar_edicion_manual(conn, cluster_id, {"email": ""}, config)

    fila = conn.execute(
        "SELECT emails_json FROM ediciones_manuales WHERE cluster_id = ?", (cluster_id,)
    ).fetchone()
    assert fila["emails_json"] is None


def test_deduplicar_retoma_corrida_incompleta_sin_recalcular(tmp_path):
    """Regresión: dos corridas reales se cortaron a mitad de camino (la
    máquina se reinició) y se perdió TODO el trabajo hecho hasta ese punto,
    porque antes solo se comiteaba al terminar. Simula una corrida
    interrumpida (decisiones_log con un corrida_id que nunca llegó a
    materializar clusters) y confirma que deduplicar_todo() la retoma:
    NO vuelve a calcular ese par (usa la decisión ya guardada, aunque sea
    distinta de lo que la regla calcularía fresca -- así se prueba que de
    verdad la está reusando y no recalculando por casualidad)."""
    from datetime import datetime, timezone

    from motor.dedup.blocking import generar_candidatos

    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "uno.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517\nJ,P,3743504517\n", encoding="utf-8"
    )
    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)

    (id_a, id_b) = next(iter(generar_candidatos(conn, config.dedup.tope_bucket)))
    corrida_vieja = "corrida-interrumpida-de-prueba"
    conn.execute(
        "INSERT INTO decisiones_log "
        "(cluster_id, raw_record_id_a, raw_record_id_b, accion, decidido_por, confianza, detalle, corrida_id, creado_en) "
        "VALUES (?, ?, ?, 'revision_pendiente', 'pendiente', NULL, 'forzado-en-el-test', ?, ?)",
        (f"pair-{id_a}-{id_b}", id_a, id_b, corrida_vieja, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    # Sin esto, un teléfono idéntico fusionaría por regla -- si el resultado
    # sigue siendo "revision_pendiente" es porque reusó la decisión vieja.

    resultado = deduplicar_todo(config, conn)

    assert resultado["revision_pendiente"] == 1
    assert resultado.get("regla", 0) == 0
    fila = conn.execute(
        "SELECT COUNT(*) AS c FROM decisiones_log WHERE corrida_id = ? AND raw_record_id_a = ? AND raw_record_id_b = ?",
        (corrida_vieja, id_a, id_b),
    ).fetchone()
    assert fila["c"] == 1  # no se logueó una segunda vez, se reusó la existente
    assert conn.execute("SELECT COUNT(*) FROM clusters WHERE corrida_id = ?", (corrida_vieja,)).fetchone()[0] == 2


def test_deduplicar_continuar_false_ignora_corrida_incompleta(tmp_path):
    from datetime import datetime, timezone

    from motor.dedup.blocking import generar_candidatos
    from motor.dedup.merge_engine import deduplicar_todo as _deduplicar_todo

    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "uno.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517\nJ,P,3743504517\n", encoding="utf-8"
    )
    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)

    (id_a, id_b) = next(iter(generar_candidatos(conn, config.dedup.tope_bucket)))
    conn.execute(
        "INSERT INTO decisiones_log "
        "(cluster_id, raw_record_id_a, raw_record_id_b, accion, decidido_por, confianza, detalle, corrida_id, creado_en) "
        "VALUES (?, ?, ?, 'revision_pendiente', 'pendiente', NULL, 'forzado-en-el-test', 'corrida-vieja', ?)",
        (f"pair-{id_a}-{id_b}", id_a, id_b, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()

    resultado = _deduplicar_todo(config, conn, continuar=False)

    assert resultado["regla"] == 1  # recalculó fresco, mismo teléfono fusiona por regla
    assert resultado.get("revision_pendiente", 0) == 0


def test_exportar_whatsapp_csv_formato_mejorows(tmp_path):
    # MejoraWS (C:\Github\Herramientas\MejoraWS) espera exactamente estas
    # 3 columnas, teléfono en E.164 SIN el "+" -- así lo pide su propio
    # README ("código de país, sin espacios ni signos").
    config = _config_prueba(tmp_path)
    (config.rutas.carpeta_raiz / "dos.csv").write_text(
        "Nombre,Apellido,Telefono\nJuan,Perez,3743504517;3743111222\n",
        encoding="utf-8",
    )
    conn = conectar(config.rutas.base_sqlite)
    extraer_todo(config, conn)
    normalizar_todo(config, conn)
    deduplicar_todo(config, conn)

    destino = exportar_whatsapp_csv(config, conn)
    assert destino.name == "contactos-whatsapp.csv"

    contenido = destino.read_text(encoding="utf-8")
    lineas = [l for l in contenido.splitlines() if l]
    assert lineas[0] == "nombre,telefono,variable"
    assert len(lineas) == 3  # un contacto, dos whatsapp -> dos filas
    for linea in lineas[1:]:
        assert "+" not in linea.split(",")[1]  # sin el "+", formato MejoraWS
        assert linea.startswith("Juan Perez,549")
